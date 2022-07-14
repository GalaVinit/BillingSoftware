from email.policy import default
from msilib import Table
from string import whitespace
from tkinter import *
from tkinter import ttk
from turtle import bgcolor, left, width
from datetime import date
from typing_extensions import final
from click import command
import tkinter.messagebox as tmsg
from PIL import Image,ImageTk
from matplotlib import image
from matplotlib.pyplot import text
from numpy import double, expand_dims, pad
from openpyxl import Workbook
from pyparsing import White

import openpyxl as xl
import openpyxl;
import os
from win32com import client

count=1
total_tamt=0
total_amt=0
total_netwt=0
total_grosswt=0
total_pcs=0

temp_pcs=0
temp_grosswt=0
temp_netwt=0
temp_amt=0
temp_tamt=0

sr=0
def clear_entries():
    pc_entry.delete(0,END)
    purity_entry.delete(0,END)
    gross_entry.delete(0,END)
    net_entry.delete(0,END)
    mk_entry.delete(0,END)
    rate_entry.delete(0,END)
    # amt_entry.delete(0,END)
    # totalamt_entry.delete(0,END)

def calcNetAmount():
    dis=discount_value.get()
    gst=sgst.get()+cgst.get()
    tamt=total_tamt
    value=tamt+(tamt*gst/100)
    return value-dis

def get_entries():
    global count
    d=descvalue.get()
    pc=pc_value.get()
    pr=purity_value.get()
    gw=gross_value.get()
    nw=net_value.get()
    mk=mk_value.get()
    rate=rate_value.get()
    amt=nw*rate
    tamt=amt+(nw*mk)
    return count,d,pc,pr,gw,nw,mk,rate,amt,tamt

def addvalues(table):
    global count
    global total_tamt
    global total_amt
    global total_netwt
    global total_grosswt
    global total_pcs

    entry=get_entries()
    count+=1
    total_tamt+=float(entry[9])
    total_amt+=float(entry[8])
    total_netwt+=float(entry[5])
    total_grosswt+=float(entry[4])
    total_pcs+=int(entry[2])

    #entry=(str(count),d,str(pc),str(pr),str(gw),str(nw),str(mk),str(rate),str(amt),str(tamt))
    table.insert(parent='',index='end',iid=count,text="Parent",values=entry)
    
    for x in finaltable.get_children():
        values=("","Total",str(total_pcs),"",str(total_grosswt),str(total_netwt),"","",str(total_amt),str(total_tamt))
        finaltable.item(x,text="",values=values)
    
    taxable_label.config(text=total_tamt)
    net_amt_label.config(text=calcNetAmount())
    clear_entries()
    #table.insert(parent='',index='end',iid=0,text="Parent",values=("1","Chain","1","22","10","10","800","4700","50000","58000"))

def select_record(table):
    global sr
    clear_entries()
    selected=table.focus()
    values=table.item(selected,'values')
    global temp_pcs
    global temp_grosswt
    global temp_netwt
    global temp_amt
    global temp_tamt

    temp_pcs=int(values[2])
    temp_grosswt=float(values[4])
    temp_netwt=float(values[5])
    temp_amt=float(values[8])
    temp_tamt=float(values[9])

    #tmsg.showinfo(values)
    sr=values[0]
    descvalue.set(values[1])
    pc_entry.insert(0,values[2])
    purity_entry.insert(0,values[3])
    gross_entry.insert(0,values[4])
    net_entry.insert(0,values[5])
    mk_entry.insert(0,values[6])
    rate_entry.insert(0,values[7])
    # amt_entry.insert(0,values[8])
    # totalamt_entry.insert(0,values[9])

def update_record(table):
    global temp_pcs
    global temp_grosswt
    global temp_netwt
    global temp_amt
    global temp_tamt

    global total_tamt
    global total_amt
    global total_netwt
    global total_grosswt
    global total_pcs

    selected=table.focus()
    #tmsg.showinfo(selected)
    values=get_entries()
    #values[0]=sr
    total_pcs=total_pcs-temp_pcs+int(values[2])
    total_grosswt=total_grosswt-temp_grosswt+int(values[4])
    total_netwt=total_netwt-temp_netwt+int(values[5])
    total_amt=total_amt-temp_amt+int(values[8])
    total_tamt=total_tamt-temp_tamt+int(values[9])

    fvalues=(sr,values[1],values[2],values[3],values[4],values[5],values[6],values[7],values[8],values[9])
    table.item(selected,text="",values=fvalues)


    for x in finaltable.get_children():
        values=("","Total",str(total_pcs),"",str(total_grosswt),str(total_netwt),"","",str(total_amt),str(total_tamt))
        finaltable.item(x,text="",values=values)
    taxable_label.config(text=total_tamt)
    net_amt_label.config(text=calcNetAmount())

def delete_record(table): 
    global total_tamt
    global total_amt
    global total_netwt
    global total_grosswt
    global total_pcs

    global count
    count-=1
    selected=table.selection()[0]
    for i in table.get_children():
        values=table.item(i,'values')
        if(values[0]>=selected):
            lst=list(values)
            lst[0]=int(values[0])-1
            lst[0]=str(lst[0])
            table.item(i,text="",values=tuple(lst))
    
    values=table.item(selected,'values')
    total_pcs=total_pcs-int(values[2])
    total_grosswt=total_grosswt-int(values[4])
    total_netwt=total_netwt-int(values[5])
    total_amt=total_amt-int(values[8])
    total_tamt=total_tamt-int(values[9])
    for x in finaltable.get_children():
        values=("","Total",str(total_pcs),"",str(total_grosswt),str(total_netwt),"","",str(total_amt),str(total_tamt))
        finaltable.item(x,text="",values=values)
    
    taxable_label.config(text=total_tamt)
    net_amt_label.config(text=calcNetAmount())
    table.delete(selected)

def print():
    sv=silver_value.get()
    gv=gold_value.get()
    cv=credit_value.get()
    dv=debit_value.get()
    cash=cash_value.get()
    if(calcNetAmount()!=sv+gv+cv+dv+cash):
        tmsg.showinfo("Error","The Net amount doesn't match the amount recieved")
        cash_entry.focus()
    else:
        source="E:/GUI/Billing/Bill.xlsx"
        wb=xl.load_workbook(source)
        ws=wb.worksheets[0]

        ws["C10"]=nvalue.get()
        ws["C11"]=addvalue.get()
        ws["C14"]=phonevalue.get()
        
        count=1
        for x in table.get_children():
            values=table.item(x,'values')
            ws["B"+str(20+count)]=values[0]
            ws["C"+str(20+count)]=values[1]
            ws["D"+str(20+count)]=values[2]
            ws["E"+str(20+count)]=values[3]
            ws["F"+str(20+count)]=values[4]
            ws["G"+str(20+count)]=values[5]
            ws["H"+str(20+count)]=values[6]
            ws["I"+str(20+count)]=values[7]
            ws["J"+str(20+count)]=values[8]
            ws["K"+str(20+count)]=values[9]
            count+=1
        ws["D31"]=total_pcs
        ws["F31"]=total_grosswt
        ws["G31"]=total_netwt
        ws["J31"]=total_amt
        ws["K31"]=total_tamt
        
        ws["D34"]=cash_value.get()
        ws["D35"]=credit_value.get()
        ws["D36"]=debit_value.get()

        ws["H34"]=gold_wt.get()
        ws["H35"]=gold_value.get()

        ws["K33"]=discount_value.get()
        ws["K34"]=total_tamt
        ws["K40"]=calcNetAmount()
        wb.save(source)
        os.startfile(source,'print')
        # img=openpyxl.drawing.image.Image("E:\GUI\Billing\LOGO.png")
        # img.width=100*2
        # img.height=100*2

    


root=Tk()
root.geometry("1000x1000")
root.config(bg="white")
root.title("Bhoomi Jewellers")

# root.wm_attributes('-transparentcolor', 'white')
main=Frame(root,padx=0)
main.pack(fill=X)
main.config(bg="white")
# photo=Image.open("images.jfif")
# resized_photo=photo.resize((1000,1000))
# bg=ImageTk.PhotoImage(resized_photo)
# Label(main,image=bg).place(x=0,y=0)

# scroll=Scrollbar(root,orient=VERTICAL)
# scroll.pack(side=RIGHT,fill=Y)
# root.configure(yscrollcommand=scroll.set)
# scroll.config(command=root.yview)






# canvas.pack(side=LEFT,fill=BOTH,expand=1)
# scroll=Scrollbar(main,orient=VERTICAL,command=canvas.yview)
# scroll.pack(side=RIGHT,fill=Y)
# canvas.configure(yscrollcommand=scroll.set)
# canvas.bind('<Configure>',lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

#           LOGO
# logo=PhotoImage(file="LOGO.png")
# logoframe=Frame(main,bg="white")
# logolabel=Label(logoframe,image=logo)
# logolabel.pack()
# logoframe.pack(anchor="n")

nvalue=StringVar()
addvalue=StringVar()
phonevalue=StringVar()
cash_value=DoubleVar()
credit_value=DoubleVar()
debit_value=DoubleVar()

gold_wt=DoubleVar()
silver_wt=IntVar()
gold_value=IntVar()
silver_value=IntVar()
discount_value=DoubleVar()
cgst=DoubleVar()
sgst=DoubleVar()

discount_value.set(0)
cgst.set(1.5)
sgst.set(1.5)

gross_amt=DoubleVar()


# Top Frame
custdet_frame=Frame(main,bg="white",padx=10,pady=10,highlightbackground="black",highlightthickness=4,bd=6,relief=SUNKEN)

name_GST_frame=Frame(custdet_frame,bg="white",pady=5)
Label(name_GST_frame,text="Name\t: ", font="Times 14 bold", bg="white").pack(anchor="nw",side=LEFT)
nentry=Entry(name_GST_frame,textvariable=nvalue,width=40)
nentry.pack(side=LEFT)
Label(name_GST_frame,text="24ADFPS0182B1ZZ", font="Times 14 italic", bg="white").pack(side=RIGHT)
Label(name_GST_frame,text="GSTIN No.\t:",font="Times 14  bold",bg="white").pack(side=RIGHT)
name_GST_frame.pack(fill=X)

add_date_frame=Frame(custdet_frame,bg="white",pady=5)
Label(add_date_frame,text="Address\t: ", font="Times 14 bold", bg="white").pack(anchor="nw",side=LEFT)
addentry=Entry(add_date_frame,textvariable=addvalue,width=40)
addentry.pack(side=LEFT)
Label(add_date_frame,text=f"{date.today()}", font="Times 14 italic", bg="white").pack(side=RIGHT)
Label(add_date_frame,text="Date\t:",font="Times 14  bold",bg="white").pack(side=RIGHT)
add_date_frame.pack(fill=X)

phone_add_frame=Frame(custdet_frame,bg="white",pady=5)
Label(phone_add_frame,text="Phone\t: ", font="Times 14 bold", bg="white").pack(anchor="nw",side=LEFT)
phoneentry=Entry(phone_add_frame,textvariable=phonevalue,width=40)
phoneentry.pack(side=LEFT)
Label(phone_add_frame,text="Bhoomi Jewellers,\n\tM. G. Road,\n\tValsad", font="Times 14 italic", bg="white").pack(side=RIGHT)
Label(phone_add_frame,text="Address\t:",font="Times 14  bold",bg="white").pack(side=RIGHT)
phone_add_frame.pack(fill=X)
custdet_frame.pack(anchor="w",pady=10,padx=20,fill=X)



#Values for items
descvalue=StringVar()
descvalue.set("NA      ")
items=["Gold Ring","Gold Chain","Earings"]

pc_value=IntVar()
purity_value=IntVar()
gross_value=DoubleVar()
net_value=DoubleVar()
rate_value=DoubleVar()
amt_value=DoubleVar()
mk_value=DoubleVar()
totalamt_value=DoubleVar()

pur_frame=Frame(main,bg="white")
pur_frame.pack(fill=X)


#dpp_frame: description pieces purity
dppgn_frame=Frame(pur_frame,bg="white",padx=15,pady=10)
dppgn_frame.pack(fill=X)
# Label(dpp_frame,text="Description: ",font="Times 14 bold",bg="white").pack(anchor="nw",side=LEFT)
Label(dppgn_frame,text="Description: ",font="Times 14 bold",bg="white").grid(row=0,column=0)
desc_menu=OptionMenu(dppgn_frame,descvalue,*items)
menu=root.nametowidget(desc_menu.menuname)
menu.config(font="Times 12 bold")
desc_menu.config(font="Times 14 bold", bg="white",width=8)
# desc_menu.pack(side=LEFT,padx=30)
desc_menu.grid(row=0,column=1,padx=(20,0))

# pc_frame=Frame(pur_frame,bg="white",padx=15,pady=10)
# pc_frame.pack(fill=X)

# Label(dpp_frame,text="Pcs: ",font="Times 14 bold",bg="white",padx=50).pack(anchor=CENTER,side="left")
Label(dppgn_frame,text="Pcs: ",font="Times 14 bold",bg="white").grid(row=0,column=2,padx=(50,0))
pc_entry=Entry(dppgn_frame,textvariable=pc_value)
pc_entry.grid(row=0,column=3)
# pc_entry.pack(side=LEFT)

# purity_frame=Frame(pur_frame,bg="white",padx=15,pady=10)
# purity_frame.pack(fill=X)
purity_entry=Entry(dppgn_frame,textvariable=purity_value)
# purity_entry.pack(anchor="e",side=RIGHT)
purity_entry.grid(row=0,column=5)
# Label(dpp_frame,text="Purity: ",font="Times 14 bold",bg="white",padx=30).pack(anchor="e",side=RIGHT)
Label(dppgn_frame,text="Purity: ",font="Times 14 bold",bg="white",padx=20).grid(row=0,column=4,padx=(50,0))



# wm_frame=Frame(pur_frame,bg="white",padx=15,pady=10)
# wm_frame.pack(fill=X)
Label(dppgn_frame,text="Gross Weight: ",font="Times 14 bold",bg="white").grid(row=0,column=6,padx=(50,0))
gross_entry=Entry(dppgn_frame,textvariable=gross_value)
gross_entry.grid(row=0,column=7)

Label(dppgn_frame,text="Net Weight: ",font="Times 14 bold",bg="white",padx=20).grid(row=0,column=8,padx=(50,0))
net_entry=Entry(dppgn_frame,textvariable=net_value)
net_entry.grid(row=0,column=9)

# mk_frame=Frame(pur_frame,bg="white",padx=15,pady=10)
# mk_frame.pack(fill=X)
mk_entry=Entry(dppgn_frame,textvariable=mk_value)
mk_entry.grid(row=1,column=1,pady=(10,0))
Label(dppgn_frame,text="Making Charges per gram: ",font="Times 14 bold",bg="white").grid(row=1,column=0,pady=(10,0))


#
# ramt_frame=Frame(pur_frame,bg="white",padx=15,pady=10)
# ramt_frame.pack(fill=X)
Label(dppgn_frame,text="Rate: ",font="Times 14 bold",bg="white").grid(row=1,column=2,pady=(10,0))
rate_entry=Entry(dppgn_frame,textvariable=rate_value)
rate_entry.grid(row=1,column=3,pady=(10,0))

# Label(ramt_frame,text="Amount: ",font="Times 14 bold",bg="white").pack(anchor="nw",side=LEFT,padx=30)
# amt_entry=Entry(ramt_frame,textvariable=amt_value)
# amt_entry.pack(anchor="w",side=LEFT)

# totalamt_entry=Entry(ramt_frame,textvariable=totalamt_value)
# totalamt_entry.pack(anchor="w",side=RIGHT)
# Label(ramt_frame,text="Total Amount: ",font="Times 14 bold",bg="white").pack(anchor="nw",side=RIGHT)

add=Button(main,text="Add Item",command=lambda: addvalues(table),bg="blue")
add.pack(pady=5)


table=ttk.Treeview(main,height=6)
table['columns']=("Sr No.","Description","Pcs","Purity","Gross Weight","Net Weight","Making Charges","Rate","Amount","Total Amount")

table.column("#0",width=0,stretch=NO)
table.column("Sr No.",anchor=CENTER,width=40)
table.column("Description",anchor=CENTER,width=220)
table.column("Pcs",anchor=CENTER,width=40)
table.column("Purity",anchor=CENTER,width=40)
table.column("Gross Weight",anchor=CENTER,width=100)
table.column("Net Weight",anchor=CENTER,width=100)
table.column("Making Charges",anchor=CENTER,width=100)
table.column("Rate",anchor=CENTER,width=100)
table.column("Amount",anchor=CENTER,width=100)
table.column("Total Amount",anchor=CENTER,width=100)

table.heading("#0",text="Label")
table.heading("Sr No.",text="Sr No.")
table.heading("Description",text="Description")
table.heading("Pcs",text="Pcs")
table.heading("Purity",text="Purity")
table.heading("Gross Weight",text="Gross Weight")
table.heading("Net Weight",text="Net Weight")
table.heading("Making Charges",text="Making Charges")
table.heading("Rate",text="Rate")
table.heading("Amount",text="Amount")
table.heading("Total Amount",text="Total Amount")

#table.insert(parent='',index='end',iid=0,text="Parent",values=("1","Chain","1","22","10","10","800","4700","50000","58000"))
table.pack()

finaltable=ttk.Treeview(main,height=1,show="tree")
finaltable['columns']=("Sr No.","Description","Pcs","Purity","Gross Weight","Net Weight","Making Charges","Rate","Amount","Total Amount")

finaltable.column("#0",width=0,stretch=NO)
finaltable.column("Sr No.",anchor=CENTER,width=40)
finaltable.column("Description",anchor=CENTER,width=220)
finaltable.column("Pcs",anchor=CENTER,width=40)
finaltable.column("Purity",anchor=CENTER,width=40)
finaltable.column("Gross Weight",anchor=CENTER,width=100)
finaltable.column("Net Weight",anchor=CENTER,width=100)
finaltable.column("Making Charges",anchor=CENTER,width=100)
finaltable.column("Rate",anchor=CENTER,width=100)
finaltable.column("Amount",anchor=CENTER,width=100)
finaltable.column("Total Amount",anchor=CENTER,width=100)

finaltable.insert(parent='',index='end',iid=0,text="Parent",values=("","Total","0","","0","0","","","0","0"))

finaltable.pack()

edit_frame=Frame(main,bg="white")
edit_frame.pack(fill=X,padx=200,pady=10)

select_button=Button(edit_frame,text="Select Record",command=lambda:select_record(table))
select_button.pack(anchor=W,side=LEFT)

update_button=Button(edit_frame,text="Update Record",command=lambda:update_record(table))
update_button.pack(anchor=E,side=RIGHT)

delete_button=Button(edit_frame,text="Delete Button", command=lambda:delete_record(table))
delete_button.pack(anchor=CENTER)

bill_frame=Frame(main,bg="white")
bill_frame.pack(fill=X,padx=20,side=TOP)

cash_reciept_frame=Frame(bill_frame,bg="white",relief=SUNKEN,bd=5,padx=20)
cash_reciept_frame.grid(row=0,column=0,columnspan=5,padx=(40,30))
#,pady=(0,90)

Label(cash_reciept_frame,text="Cash Reciept",font="Times 14 bold",bg="white").grid(row=0,columnspan=2)
Label(cash_reciept_frame,text="Cash: \t",font="Times 14 bold", bg="white").grid(row=1,column=0,padx=(0,15))

cash_entry=Entry(cash_reciept_frame,textvariable=cash_value,width=30)
cash_entry.grid(row=1,column=1)

Label(cash_reciept_frame,text="Credit Card: \t",font="Times 14 bold", bg="white").grid(row=2,column=0,padx=(0,15))
credit_entry=Entry(cash_reciept_frame,textvariable=credit_value,width=30)
credit_entry.grid(row=2,column=1)

Label(cash_reciept_frame,text="Debit Card: \t",font="Times 14 bold", bg="white").grid(row=3,column=0,padx=(0,15))
credit_entry=Entry(cash_reciept_frame,textvariable=debit_value,width=30)
credit_entry.grid(row=3,column=1)

Label(cash_reciept_frame,text="  ",bg="white").grid(row=4,column=0)
Label(cash_reciept_frame,text="  ",bg="white").grid(row=4,column=1)

old_frame=Frame(bill_frame,bg="white",relief=SUNKEN,bd=5,padx=20)
old_frame.grid(row=0,column=7,columnspan=5,padx=(70,20))
# ,pady=(0,60),padx=(10,10)
# ,pady=10

Label(old_frame,text="Metal Entry",font="Times 14 bold",bg="white").grid(row=0,columnspan=2)

Label(old_frame,text="Old-Gold wt:",font="Times 14 bold", bg="white").grid(row=1,column=0,padx=(0,15))
gold_wt_entry=Entry(old_frame,textvariable=gold_wt,width=30)
gold_wt_entry.grid(row=1,column=1)

Label(old_frame,text="Old-Gold value:",font="Times 14 bold", bg="white").grid(row=2,column=0,padx=(0,15))
gold_value_entry=Entry(old_frame,textvariable=gold_value,width=30)
gold_value_entry.grid(row=2,column=1)

Label(old_frame,text="Old-Silver wt:",font="Times 14 bold", bg="white").grid(row=3,column=0,padx=(0,15))
silver_wt_entry=Entry(old_frame,textvariable=silver_wt,width=30)
silver_wt_entry.grid(row=3,column=1)

Label(old_frame,text="Old-Silver value:",font="Times 14 bold", bg="white").grid(row=4,column=0,padx=(0,15))
silver_value_entry=Entry(old_frame,textvariable=silver_value,width=30)
silver_value_entry.grid(row=4,column=1)

final_amt_frame=Frame(bill_frame,bg="white",relief=SUNKEN,bd=5,padx=20)
final_amt_frame.grid(row=0,column=13,columnspan=5,padx=(70,20))
# ,pady=(0,60),padx=(10,10)
# ,pady=10

Label(final_amt_frame,text="Taxable amount:",font="Times 14 bold", bg="white").grid(row=1,column=0,padx=(0,15))
taxable_label=Label(final_amt_frame,text="",font="Times 14 bold", bg="white")
taxable_label.grid(row=1,column=1,padx=(0,15))

Label(final_amt_frame,text="Discount:",font="Times 14 bold", bg="white").grid(row=2,column=0,padx=(0,15))
discount_entry=Entry(final_amt_frame,textvariable=discount_value,width=30)
discount_entry.grid(row=2,column=1)

Label(final_amt_frame,text="CGST:",font="Times 14 bold", bg="white").grid(row=3,column=0,padx=(0,15))
cgst_entry=Entry(final_amt_frame,textvariable=cgst,width=30)
cgst_entry.grid(row=3,column=1)

Label(final_amt_frame,text="SGST:",font="Times 14 bold", bg="white").grid(row=4,column=0,padx=(0,15))
sgst_entry=Entry(final_amt_frame,textvariable=sgst,width=30)
sgst_entry.grid(row=4,column=1)

Label(final_amt_frame,text="Net Amount:",font="Times 14 bold", bg="white").grid(row=5,column=0,padx=(0,15))
net_amt_label=Label(final_amt_frame,text="",font="Times 14 bold", bg="white")
net_amt_label.grid(row=5,column=1)

# print_frame=Frame(main,bg="white")
# print_frame.pack(side=LEFT)
print_button=Button(main,text="Print",font="Times 14 bold",command=print,height=1)
print_button.pack(anchor=CENTER,pady=10)

root.mainloop()
