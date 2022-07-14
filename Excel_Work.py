import openpyxl as xl
import openpyxl;
'''
source="E:/GUI/Billing/Bill.xlsx"
wb1=xl.load_workbook(source)
ws1=wb1.worksheets[0]

destination="E:/GUI/Billing/file1.xlsx"
wb2=xl.load_workbook(destination)
ws2=wb2.active

mr=ws1.max_row
mc=ws1.max_column

for i in range(1,mr+1):
    for j in range(1,mc+1):
        c=ws1.cell(row=i,column=j)
        ws2.cell(row=i,column=j).value=c.value

ws2.merge_cells('D1:J11')
img=openpyxl.drawing.image.Image("E:\GUI\Billing\LOGO.png")
# img.width=25*1
# img.height=25*1
ws2.add_image(img,'D1')


wb2.save(str(destination))
'''
wb1=xl.load_workbook("E:\GUI\Billing\Bill.xlsx")




# from openpyxl import Workbook, load_workbook
# workbook = load_workbook(filename="E:\GUI\Billing\Bill.xlsx")

# sheet = workbook.active

# sheet["A1"]="Full Name"
# sheet["A2"]="Hi"

# workbook.save(filename="E:\GUI\Billing\Bill.xlsx")

# f=open("file2.xlsx","w")


# import xlwt
# import xlrd
# from xlutils.copy import copy

# # load the excel file
# rb = xlrd.open_workbook('Bill.xlsx')

# # copy the contents of excel file
# wb = copy(rb)

# # open the first sheet
# w_sheet = wb.get_sheet(0)

# # row number = 0 , column number = 1
# w_sheet.write(0,1,'Modified !')

# # save the file
# wb.save('UserBook.xls')
